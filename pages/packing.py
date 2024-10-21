import logging
import os
import time

from openpyxl.reader.excel import load_workbook
from win32ctypes.pywin32 import pywintypes

from utils.app_core import ManufacturePage
from utils.models import MattressRequest
from utils.tools import config, print_file, get_date_str
import streamlit as st


def form_box_text(task):
    # Текст контейнера красится в красный, когда у наряда приоритет
    return ((':red' if task.high_priority else ':gray') +
            f"[**Артикул:** {task.article}  \n" +
            f"**Размер**: {task.size}  \n" +
            f"**Топ:** {task.base_fabric}  \n" +
            f"**Бок:** {task.side_fabric}  \n" +
            f"**ПБ:** {task.springs}  \n" +
            (f"**Комментарий**: {task.comment}  " if task.comment else '') + "]")


class PackingPage(ManufacturePage):
    def __init__(self, name, icon):
        super().__init__(name, icon)
        self.default_printer_name = config.get('site').get('hardware').get('default_printer')
        self.label_printer_name = config.get('site').get('hardware').get('label_printer')

    def talon_button(self, order, task):
        if st.button(label=f":blue[**Талон**]", key=f"print_talon_button_{task.id}"):
            template_path = 'static/guarantee.xlsx'
            wb = load_workbook(template_path)
            ws = wb.active

            # Заполнение шаблона
            ws['B4'] = "Матрас АРТ.№ " + task.article + '  |  ПБ: ' + task.springs
            ws['B6'] = task.size
            ws['B8'] = task.deadline
            ws['B16'] = f"{order.organization} - {order.address}"

            document_path = fr'cash\{self.page_name}_talon_{order.id}.xlsx'
            wb.save(document_path)
            try:
                print_file(document_path, self.default_printer_name)
                st.toast("Печать талона...", icon='🖨️')
                time.sleep(1)
            except pywintypes.error as e:
                st.toast(f"Ошибка печати: {e}")
            try:
                os.remove(document_path)
            except Exception:
                time.sleep(0.5)
                os.remove(document_path)

    def label_button(self, task):

        if st.button(label=f":orange[**Этикетка**]", key=f"print_label_button_{task.id}"):
            try:
                label_pdf_path = fr"static\labels\{task.article}.pdf"
                printer = self.label_printer_name

                print_file(label_pdf_path, printer)

                st.toast("Печать этикетки...", icon='🖨️')

            except FileNotFoundError:
                st.toast("Ошибка печати. Шаблон для этикетки не найден.", icon='❗')

            except Exception as e:
                st.toast(f"Ошибка печати: {e}")

    def tasks_tiles(self, order, num_columns: int = 3):
        """Принимает отфильтрованные данные. Выводит заявки в виде плиточек на страницу.
        Отфильтрованные данные выводятся, а потом, при нажатии "Готово" на заявке, по
    индексу изменения соотносятся с главным хранилищем и записываются"""
        row_container = st.columns(num_columns)

        count = 0

        for task in order.mattress_requests:

            if (not task.components_is_done or
                not task.fabric_is_done or
                not task.gluing_is_done or
                not task.sewing_is_done or
                    task.packing_is_done):
                continue

            if count % num_columns == 0:
                row_container = st.columns(num_columns)

            box = row_container[count % num_columns].container(border=True)
            with box:
                main_row, photo_space, buffer = st.columns([18, 1, 1])
                photo = task.photo

                if photo:
                    main_row, photo_space, buffer = st.columns([10, 2, 2])
                    with photo_space:
                        st.image(photo, caption='Фото', width=80)
                with main_row:
                    st.markdown(form_box_text(task))

                button_row_1, button_row_2, button_row_3 = st.columns([1, 1, 1])

                with button_row_1:
                    if st.button(f":green[**{self.done_button_text}**]", key=f'button_packing_done_{task.id}'):
                        history_note = f'{self.pages_history_note()} \n'
                        db_task = self.session.query(MattressRequest).get(task.id)

                        if db_task:
                            db_task.packing_is_done = True
                            db_task.history += history_note
                            self.update_db(db_task)
                        else:
                            logging.error(f"Task with id {task.id} not found in the database")
                        st.rerun()

                with button_row_2:
                    self.talon_button(order, task)

                with button_row_3:
                    self.label_button(task)

            count += 1

    @st.fragment(run_every=3)
    def tiles_rows(self):

        employee = st.session_state.get(self.page_name)
        if not employee:
            st.warning("Сначала отметьте сотрудника.")
            return

        orders = self.get_orders_with_mattress_requests()
        if not orders:
            st.info("Пока нечего упаковывать.")
            return

        for order in orders:
            # Проверяем, есть ли активные заявки на матрасы
            has_active_requests = any((request.sewing_is_done and
                                       not request.packing_is_done) for request in order.mattress_requests)

            if not has_active_requests:
                continue

            contact = order.organization or order.contact or 'Клиент'
            delivery_type = order.delivery_type or 'Самовывоз'
            region = order.region if order.delivery_type != 'Самовывоз' else ''
            address = order.address or 'Цех'

            st.markdown(f"#### {region} {delivery_type}, {contact}, {address}")
            with st.expander(f"№{order.id}: {get_date_str(order.created)}", expanded=True):
                self.tasks_tiles(order, 4)


Page = PackingPage("Упаковка", "📦")
Page.tiles_rows()
