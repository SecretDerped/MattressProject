import datetime
import os
import time

import streamlit as st
from openpyxl.reader.excel import load_workbook

from utils.tools import config, read_file, side_eval, get_date_str, save_to_file, time_now, create_cashfile_if_empty, \
    load_tasks, print_file


class Page:
    def __init__(self, page_name, icon):
        self.page_name = page_name
        self.icon = icon

        self.employees_cash = config.get('site').get('employees_cash_filepath')
        self.employee_columns_config = {
            "is_on_shift": st.column_config.CheckboxColumn("На смене", default=False),
            "name": st.column_config.TextColumn("Имя / Фамилия", default=''),
            "position": st.column_config.TextColumn("Роли", width='medium', default=''),
            "barcode": st.column_config.LinkColumn("Штрих-код", display_text="Открыть", disabled=True),
        }
        create_cashfile_if_empty(self.employee_columns_config, self.employees_cash)

        self.task_cash = config.get('site').get('tasks_cash_filepath')
        self.tasks_columns_config = {
            "high_priority": st.column_config.CheckboxColumn("Приоритет", default=False),
            "deadline": st.column_config.DateColumn("Срок",
                                                    format="DD.MM.YYYY",
                                                    step=1,
                                                    default=datetime.date.today()),
            "article": "Артикул",
            "size": "Размер",
            "base_fabric": st.column_config.TextColumn("Ткань (Верх / Низ)",
                                                       default='Текстиль'),
            "side_fabric": st.column_config.TextColumn("Ткань (Бок)",
                                                       default='Текстиль'),
            "photo": st.column_config.ImageColumn("Фото", help="Кликните, чтобы развернуть"),
            "comment": st.column_config.TextColumn("Комментарий",
                                                   default='',
                                                   width='small'),
            "springs": st.column_config.TextColumn("Пружины",
                                                   default=''),
            "attributes": st.column_config.TextColumn("Состав начинки",
                                                      default='',
                                                      width='medium'),
            "fabric_is_done": st.column_config.CheckboxColumn("Нарезано",
                                                              default=False),
            "gluing_is_done": st.column_config.CheckboxColumn("Собран",
                                                              default=False),
            "sewing_is_done": st.column_config.CheckboxColumn("Пошит",
                                                              default=False),
            "packing_is_done": st.column_config.CheckboxColumn("Упакован",
                                                               default=False),
            "history": st.column_config.TextColumn("Действия",
                                                   width='small',
                                                   disabled=True),
            "client": st.column_config.TextColumn("Заказчик",
                                                  default='',
                                                  width='medium'),
            "delivery_type": st.column_config.SelectboxColumn("Тип доставки",
                                                              options=config.get('site').get('delivery_types'),
                                                              default=config.get('site').get('delivery_types')[0],
                                                              required=True),
            "address": st.column_config.TextColumn("Адрес",
                                                   default='Наш склад',
                                                   width='large'),
            "region": st.column_config.SelectboxColumn("Регион",
                                                       width='medium',
                                                       options=config.get('site').get('regions'),
                                                       default=config.get('site').get('regions')[0],
                                                       required=True),
            "created": st.column_config.DatetimeColumn("Создано",
                                                       format="D.MM.YYYY | HH:MM",
                                                       disabled=True),
        }
        create_cashfile_if_empty(self.tasks_columns_config, self.task_cash)

        st.set_page_config(page_title=self.page_name,
                           page_icon=self.icon,
                           layout="wide")

    def header(self):
        st.title(f'{self.icon} {self.page_name}')

    @staticmethod
    def load_tasks():
        return load_tasks()


class ManufacturePage(Page):
    def __init__(self, page_name, icon):
        super().__init__(page_name, icon)
        self.talon_button_text = 'Талон'
        self.label_button_text = 'Этикетка'
        self.done_button_text = 'Готово'
        self.label_printer_name = 'Xprinter xp-370b'
        self.header()

    def header(self):
        col1, col2 = st.columns([3, 1])
        with col1:
            st.title(f'{self.icon} {self.page_name}')
        with col2:
            self.employee_choose()

    def employees_on_shift(self, searching_position: str) -> list:
        """Принимает строку для фильтрации по роли сотрудника (position).
        Читает .pkl из employees_cash_filepath, преобразует в датафрейм.
        Фильтрует записи, где значение в "is_on_shift" стоит True,
        а в колонке "position" есть подстрока из аргумента функции (независимо от регистра).

        :returns: [список имен сотрудников на смене по искомой роли]"""
        dataframe = read_file(self.employees_cash)
        if 'is_on_shift' not in dataframe.columns or 'position' not in dataframe.columns:
            raise ValueError("В датафрейме сотрудников должны быть колонки 'is_on_shift' и 'position'")

        filtered_df = dataframe[(dataframe['is_on_shift'] == True) & (
            dataframe['position'].str.contains(searching_position, case=False, na=False))]

        return filtered_df['name'].tolist()

    @st.experimental_fragment(run_every="4s")
    def employee_choose(self):
        """Виджет для выбора активного сотрудника для рабочего места.
        Рабочее место - строка в качестве аргумента.
        Показывает выпадающий список из сотрудников, находящихся на смене.
        При выборе сотрудника в session_state сохраняется строка с именем сотрудника под ключом с названием должности.
        Пример: st.session_state['швейный стол'] == 'Полиграф Полиграфович'"""

        def save_employee(position):
            """Метод для работы эффекта on_change виджета из employee_choose.
            Без него выбираемый сотрудник корректно не записывается в session_state."""
            st.session_state[position] = st.session_state[position]

        st.selectbox('Ответственный',
                     placeholder="Выберите сотрудника",
                     options=self.employees_on_shift(self.page_name),
                     index=None,
                     key=self.page_name,
                     on_change=save_employee, args=(self.page_name,))

    # Только на руинах прошлого можно построить империю будущего. Этот костыль стал надстройкой над когда-то славным
    # классом, от которого остался лишь скелет
    def talon_button(self, row, index):
        file_name = f'{self.page_name}_talon_{index}.xlsx'
        # Кнопка для печати талона

        if st.button(label=f":blue[**{self.talon_button_text}**]", key=file_name):
            template_path = 'static/guarantee.xlsx'
            wb = load_workbook(template_path)
            ws = wb.active

            # Заполнение шаблона
            ws['B4'] = "Матрас АРТ.№ " + row['article'] + '  |  ПБ: ' + row['springs']

            ws['B6'] = row['size']

            ws['B8'] = row['deadline'].strftime('%d.%m.%Y')

            ws['B16'] = f"{row['client']}  {row['address']}" if row['address'] else 'Краснодар, ул. Демуса 6А'

            file_path = fr'cash\{file_name}'
            wb.save(file_path)
            print_file(file_path)
            st.toast("Сейчас распечатается талон...", icon='🖨️')
            time.sleep(1)
            try:
                os.remove(file_path)
            except Exception:
                time.sleep(0.5)

    def label_button(self, row, index):
        article = row['article']
        if st.button(label=f":orange[**{self.label_button_text}**]", key=f"{article}_{index}"):
            try:
                file_path = fr"static\labels\{article}.pdf"
                print_file(file_path, self.label_printer_name)
            except FileNotFoundError:
                print_file("static\labels\800.pdf")

    @staticmethod
    def inner_box_text(row):
        """Метод, выдающий текст внутри бокса. Для каждой страницы с плитками заявок можно переопределять этот метод."""
        return (f"**Артикул:** {row.get('article')}  \n"
                f"**Ткань**: {row.get('base_fabric')}  \n"
                f"**Тип доставки**: {row.get('delivery_type')}  \n"
                f"**Адрес:** {row.get('address')}  \n"
                f"**Клиент:** {row.get('client')}  \n"
                f"**Верх/Низ**: {row.get('base_fabric')}  \n"
                f"**Бочина**: {row.get('side_fabric')}  \n"
                f"**Состав:** {row.get('attributes')}  \n"
                f"**Размер:** {row.get('size')} ({side_eval(row.get('size'), row.get('side_fabric'))}  \n"
                f"**Срок**: {get_date_str(row.get('deadline'))}  \n")

    def _form_box_text(self, index, row):
        # Текст контейнера красится в красный, когда у наряда приоритет
        text_color = 'red' if row['high_priority'] else 'gray'

        box_text = f':{text_color}[{self.inner_box_text(row)}'

        if row['comment']:
            box_text += f"**Комментарий**: {row['comment']}  "

        box_text += ']'
        return box_text

    def show_tasks_tiles(self, data, stage_to_done: str, num_columns: int = 3) -> bool:
        """Принимает отфильтрованные данные. Выводит заявки в виде плиточек на страницу.
        Отфильтрованные данные выводятся, а потом, при нажатии "Готово" на заявке, по
        индексу изменения соотносятся с главным хранилищем и записываются"""
        main_data = super().load_tasks()
        page = self.page_name

        if len(data) == 0:
            return st.header('Все заявки выполнены! Хорошая работа 🏖️')

        row_container = st.columns(num_columns)
        count = 0
        for index, row in data.iterrows():
            if count % num_columns == 0:
                row_container = st.columns(num_columns)

            box = row_container[count % num_columns].container(height=250, border=True)

            with box:
                photo = row['photo']
                if photo:
                    col1, col2, buff, col3 = st.columns([12, 3, 1, 6])
                    col2.image(photo, caption='Фото', width=80)
                else:
                    col1, col3 = st.columns([3, 1])

                with col1:
                    st.markdown(self._form_box_text(index, row))
                with col3:
                    if page in st.session_state and st.session_state[page]:
                        employee = st.session_state[page]

                        if st.button(f":green[**{self.done_button_text}**]", key=f'{page}_done_{index}'):
                            history_note = f'({time_now()}) {page} [ {employee} ] -> {self.done_button_text}; \n'
                            main_data.at[index, stage_to_done] = True
                            main_data.at[index, 'history'] += history_note
                            save_to_file(main_data, self.task_cash)
                            st.rerun()

                        self.talon_button(row, index)
                        self.label_button(row, index)
            count += 1
        return True
